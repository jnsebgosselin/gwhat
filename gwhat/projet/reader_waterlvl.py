# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# Copyright © GWHAT Project Contributors
# https://github.com/jnsebgosselin/gwhat
#
# This file is part of GWHAT (Ground-Water Hydrograph Analysis Toolbox).
# Licensed under the terms of the GNU General Public License.
# -----------------------------------------------------------------------------


# ---- Standard library imports
from copy import deepcopy
import re
import os
import os.path as osp
import csv
from collections import OrderedDict
from collections.abc import Mapping

# ---- Third party imports
import numpy as np
import pandas as pd
import xlrd
import openpyxl


FILE_EXTS = ['.csv', '.xls', '.xlsx']


INDEX = 'Time'

COL_REGEX = OrderedDict([
    (INDEX, r'(date|time|datetime)'),
    ('BP', r'(bp|baro|patm)'),
    ('WL', r'(wl|waterlevels)'),
    ('ET', r'(et|earthtides)')
    ])
COLUMNS = list(COL_REGEX.keys())

HEADER = {'Well': '',
          'Well ID': '',
          'Province': '',
          'Municipality': '',
          'Latitude': 0,
          'Longitude': 0,
          'Elevation': 0}
HEADER_REGEX = {
    'Well': r'(?<!\S)(well|wellname|name)(:|=)?(?!\S)',
    'Well ID': r'(?<!\S)(wellid|id)(:|=)?(?!\S)',
    'Province': r'(?<!\S)(province|prov)(:|=)?(?!\S)',
    'Municipality': r'(?<!\S)municipality(:|=)?(?!\S)',
    'Latitude': r'(?<!\S)(latitude|lat)(:|=)?(?!\S)',
    'Longitude': r'(?<!\S)(longitude|lon)(:|=)?(?!\S)',
    'Elevation': r'(?<!\S)(elevation|elev|altitude|alt)(:|=)?(?!\S)'
    }


def _format_column_names(df):
    """
    Rename valid columns, drop invalid columns, and add missing columns.
    """
    # Rename valid columns and drop invalid columns.
    drop = []
    rename = {}
    for column in df.columns:
        for colname, regex in COL_REGEX.items():
            str_ = column.replace(" ", "").replace("_", "")
            if re.search(regex, str_, re.IGNORECASE):
                rename[column] = colname
                break
        else:
            drop.append(column)
    df = df.rename(columns=rename)
    df = df.drop(columns=drop)

    # Add missing columns.
    for column in COLUMNS:
        if column not in df.columns:
            df[column] = np.nan

    return df[COLUMNS].copy()


def _format_numeric_data(df):
    """Format the data to floats type."""
    for colname in COLUMNS:
        if colname != INDEX and colname in df.columns:
            df[colname] = pd.to_numeric(df[colname], errors='coerce')
    return df


def _format_datetime_data(df):
    """Format the dates to datetimes and set it as index."""
    if INDEX not in df.columns:
        print('WARNING: no "Time" data found in the datafile.')
        return df

    if isinstance(df['Time'][0], (int, float)):
        # Time needs to be converted from Excel numeric dates
        # to ISO date strings.
        datetimes = df['Time'].astype('float64', errors='raise')
        datetimes = pd.to_datetime(
            datetimes.apply(
                lambda date: xlrd.xldate.xldate_as_datetime(date, 0)))

        # Get rid of milliseconds to avoid introducting
        # round-off errors.
        datetimes = datetimes.dt.round('S')

        df['Time'] = datetimes
    elif isinstance(df['Time'][0], (bytes)):
        strtimes = df['Time'].apply(lambda x: x.decode())
        df['Time'] = pd.to_datetime(strtimes, infer_datetime_format=True)
    elif isinstance(df['Time'][0], (str)):
        df['Time'] = pd.to_datetime(df['Time'], infer_datetime_format=True)
    else:
        print('WARNING: the dates are not formatted correctly.')

    df.set_index(['Time'], drop=True, inplace=True)
    return df


def _drop_duplicates(df):
    """
    Drop duplicated indexes from the dataframe.
    """
    if df.index.duplicated(keep='first').any():
        print("WARNING: Duplicated values were found in the datafile. "
              "Only the first entries for each date were kept.")
        index = df.index.drop_duplicates(keep='first')
        df = df.loc[index]
    return df


class WLDataFrame(pd.DataFrame):
    def __init__(self, data=None, columns=None, metadata=None):
        if data is None:
            super().__init__(data=[], columns=COLUMNS)
            self.set_index(INDEX, drop=True, inplace=True)
        else:
            df = pd.DataFrame(data, columns=columns)
            df = _format_column_names(df)
            df = _format_numeric_data(df)
            df = _format_datetime_data(df)
            df = _drop_duplicates(df)
            super().__init__(df)

        metadata = {} if metadata is None else metadata
        for key, val in HEADER.items():
            self.attrs[key] = metadata.get(key, val)


def open_water_level_datafile(filename):
    """Open a water level data file and return the data."""
    root, ext = os.path.splitext(filename)
    if ext not in FILE_EXTS:
        raise ValueError("Supported file format are: ", FILE_EXTS)
    else:
        print('Loading waterlvl time-series from "%s"...' %
              osp.basename(filename))

    if ext == '.csv':
        with open(filename, 'r', encoding='utf8') as f:
            data = list(csv.reader(f, delimiter=','))
    elif ext == '.xls':
        with xlrd.open_workbook(filename, on_demand=True) as wb:
            sheet = wb.sheet_by_index(0)
            data = [sheet.row_values(rowx, start_colx=0, end_colx=None) for
                    rowx in range(sheet.nrows)]
    elif ext == '.xlsx':
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook[workbook.sheetnames[0]]
            data = [list(row_values) for row_values in
                    sheet.iter_rows(min_col=1, values_only=True)]
        finally:
            workbook.close()
    return data


def read_water_level_datafile(filename):
    """
    Load a water level dataset from a csv or an Excel file and format the
    data in a Pandas dataframe with the dates used as index.
    """
    reader = open_water_level_datafile(filename)

    # Fetch the metadata from the header.
    header = deepcopy(HEADER)
    for i, row in enumerate(reader):
        if not len(row):
            continue

        label = str(row[0]).replace(" ", "").replace("_", "")
        if re.search(COL_REGEX[INDEX], label, re.IGNORECASE):
            break

        for key in HEADER.keys():
            if re.search(HEADER_REGEX[key], label, re.IGNORECASE):
                if isinstance(header[key], (float, int)):
                    try:
                        header[key] = float(row[1])
                    except ValueError:
                        print('Wrong format for metadata "{}".'.format(key))
                else:
                    header[key] = str(row[1])
                break
    else:
        print("ERROR: no data found in input water level file.")
        return WLDataFrame(metadata=header)

    root, ext = osp.splitext(filename)
    if ext.lower() in ['.xls', '.xlsx']:
        data = pd.read_excel(
            filename,
            header=i,
            parse_dates=[row[0]]
            )
    else:
        data = pd.read_csv(
            filename,
            skip_blank_lines=False,
            header=i)

    # Cast the data into a Pandas dataframe.
    dataf = WLDataFrame(data, columns=None, metadata=header)
    dataf.filename = filename

    return dataf


def load_waterlvl_measures(filename, well):
    """
    Load and read the water level manual measurements from the specified
    resource file for the specified well.
    """
    print('Loading manual water level measures for well %s...' % well, end=" ")
    # Determine the extension of the file.
    root, ext = os.path.splitext(filename)
    exts = [ext] if ext in FILE_EXTS else FILE_EXTS
    for ext in exts:
        filename = root + ext
        if os.path.exists(root + ext):
            break
    else:
        print("done")
        return np.array([]), np.array([])

    # Open and read the file.
    dtypes = {'Well_ID': 'str', 'Time (days)': 'float', 'Obs. (mbgs)': 'float'}
    if ext == '.csv':
        data = pd.read_csv(filename, dtype=dtypes)
    elif ext in ['.xlsx', '.xls']:
        data = pd.read_excel(filename, dtype=dtypes)

    well_data = data[data['Well_ID'] == well]
    wl_mes = well_data['Obs. (mbgs)'].values
    time_mes = well_data['Time (days)'].values
    print("done")

    return time_mes, wl_mes


def generate_HTML_table(name, lat, lon, alt, mun):

    FIELDS = [['Well Name', name],
              ['Latitude', '%0.3f°' % lat],
              ['Longitude', '%0.3f°' % lon],
              ['Altitude', '%0.1f m' % alt],
              ['Municipality', mun]]

    table = '<table border="0" cellpadding="2" cellspacing="0" align="left">'
    for row in FIELDS:
        table += '''
                 <tr>
                   <td width=10></td>
                   <td align="left">%s</td>
                   <td align="left" width=20>:</td>
                   <td align="left">%s</td>
                   </tr>
                 ''' % (row[0], row[1])
    table += '</table>'

    return table


class WLDatasetBase(Mapping):
    """
    A water level data frame base class.
    """

    def __init__(self, *args, **kwargs):
        super().__init__()
        self.dset = None
        self._undo_stack = []
        self._dataf = WLDataFrame()

    def __load_dataset__(self):
        """Loads the dataset and save it in a store."""
        raise NotImplementedError

    def __len__(self):
        return len(self._dataf)

    def __setitem__(self, key, value):
        raise NotImplementedError

    def __iter__(self):
        raise NotImplementedError

    # ---- Attributes
    @property
    def data(self):
        return self._dataf

    @property
    def xldates(self):
        """
        Return a numpy array containing the Excel numerical dates
        corresponding to the dates of the dataset.
        """
        if 'XLDATES' not in self._dataf.columns:
            print('Converting datetimes to xldates...', end=' ')
            timedeltas = (
                self._dataf.index - xlrd.xldate.xldate_as_datetime(4000, 0))
            self._dataf['XLDATES'] = (
                timedeltas.total_seconds()/(3600 * 24) + 4000)
            print('done')
        return self._dataf['XLDATES'].values

    @property
    def dates(self):
        return self.data.index.values

    @property
    def strftime(self):
        return self.data.index.strftime("%Y-%m-%dT%H:%M:%S").values.tolist()

    @property
    def waterlevels(self):
        return self.data['WL'].values

    # ---- Versionning
    @property
    def has_uncommited_changes(self):
        """"
        Return whether there is uncommited changes to the water level data.
        """
        return bool(len(self._undo_stack))

    def commit(self):
        """Commit the changes made to the water level data to the project."""
        raise NotImplementedError

    def undo(self):
        """Undo the last changes made to the water level data."""
        if self.has_uncommited_changes:
            changes = self._undo_stack.pop(-1)
            self._dataf['WL'][changes.index] = changes

    def clear_all_changes(self):
        """
        Clear all changes that were made to the water level data since the
        last commit.
        """
        while self.has_uncommited_changes:
            self.undo()

    def delete_waterlevels_at(self, indexes):
        """Delete the water level data at the specified indexes."""
        if len(indexes):
            self._add_to_undo_stack(indexes)
            self._dataf['WL'].iloc[indexes] = np.nan

    def _add_to_undo_stack(self, indexes):
        """
        Store the old water level values at the specified indexes in a stack
        before changing or deleting them. This allow to undo or cancel any
        changes made to the water level data before commiting them.
        """
        if len(indexes):
            self._undo_stack.append(self._dataf['WL'].iloc[indexes].copy())


class WLDataset(WLDatasetBase):
    """
    A water level dataset container that loads its data from a csv
    or an Excel file.
    """

    def __init__(self, filename, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.__load_dataset__(filename)

    def __getitem__(self, key):
        """Returns the value saved in the store at key."""
        if key == INDEX:
            return self.strftime
        elif key in COLUMNS:
            return self.data[key].values
        elif key in HEADER.keys():
            return self.data.attrs[key]
        elif key == 'filename':
            return self.data.filename

        return self.dset.__getitem__(key)

    def __setitem__(self, key, value):
        if key in HEADER.keys():
            self.data.attrs[key] = value
        else:
            raise KeyError(key)

    def __load_dataset__(self, filename):
        """Loads the dataset from a file and saves it in the store."""
        self._dataf = read_water_level_datafile(filename)


if __name__ == "__main__":
    from gwhat import __rootdir__
    dirname = osp.join(__rootdir__, 'projet', 'tests', 'data')
    df1 = WLDataset(osp.join(dirname, "water_level_datafile.csv"))
    df2 = WLDataset(osp.join(dirname, "water_level_datafile.xls"))
    df3 = WLDataset(osp.join(dirname, "water_level_datafile.xlsx"))
    df4 = WLDataset(osp.join(dirname, "water_level_datafile_xldates.csv"))
    df5 = WLDataset(osp.join(dirname, "water_level_datafile_strfmt.xls"))
    df6 = WLDataset(osp.join(dirname, "water_level_datafile_strfmt.xlsx"))
    df7 = WLDataFrame()
